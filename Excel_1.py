#установим соответствующие библиотеки
#pip install openpyxl

from openpyxl import load_workbook #импортировали модуль для работы с Иксель
import openpyxl

#создаем файл
wb = openpyxl.Workbook()
wb.save('МойТест_1.xlsx')

wd_form = load_workbook(filename= 'Тестовый документ.xlsx') #вытягиваем формулы

#узнать активный лист
active_sheet = wd_form.active
print(active_sheet)

wd_value = load_workbook(filename= 'Тестовый документ.xlsx', data_only=True ) #вытягиваем значения

sheet_form = wd_form['ФИО+ДР+ЗП'] #получаем формулы с конкретного листа
sheet_value = wd_value['ФИО+ДР+ЗП'] #получаем данные с конкретного листа

E6_form = sheet_form['E6'].value
E6_value = sheet_value['E6'].value

#можно получить значение ячейки через активный лист
cell_obj = active_sheet.cell(row = 6, column = 5)
print(cell_obj.value)

#получим максимальное число имеющихся заполненных строк и столбцов
print("Число заполненных строк: " + str(active_sheet.max_row))
print("Число заполненных столбцов: " + str(active_sheet.max_column))

#получим имена заполненных колонок
max_column = active_sheet.max_column

for i in range(1, max_column + 1): #потому что двигаемся ДО => плюсуем 1
    cell_obj = active_sheet.cell(row = 1, column = i)
    print(cell_obj.value)

#получим имена заполненной строки
max_row = active_sheet.max_row
for i in range(1, max_row + 1): #потому что двигаемся ДО => плюсуем 1
    cell_obj_row = active_sheet.cell(row = i, column = 5)
    print(cell_obj_row.value)


print(E6_form)
print(E6_value)