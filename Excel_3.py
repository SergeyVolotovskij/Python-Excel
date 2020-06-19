#импортируем необходимые библиотеки
from openpyxl import workbook
from openpyxl.styles import Font, Color, colors
from openpyxl import load_workbook

#для удобства указали название файла
filename = 'МойТест_3.xlsx'

#вытянули данные с документа
active_excel = load_workbook(filename=filename, data_only=True)

#делаем или смотрим активный лист
active_sheet = active_excel.active

#меняем наименование листа
active_sheet.title = 'MyGameDesign'

#нужно понять максимальный размер данных на листе

def max_min_data(x):
    if x == 'max_row':
        max_row = int(active_sheet.max_row)
        return max_row
    else:
        max_column = int(active_sheet.max_column)
        return max_column
    #print('СТРОК: ' + str(max_row) + '\nКОЛОНОК: ' + str(max_column))

#создадим список, который запишем в столбик
Spisok1 = [
    'ПОШЛЕМ ТРВ',
    1,
    2,
    3,
    4
]

Spisok2 = [
    'Фамилия',
    'Имя',
    'Отчество',
    'Год рождения',
    'Зарплата'
]

Spisok3 = [
    'Запишем',
    'все',
    'в',
    'одну',
    'строку'
]

#с помощью for запишем наши данные (по умолчанию в 1 колонке после записей
#либо с самого начала если документ пустой)
for row in range(len(Spisok1)):
    active_sheet.append([Spisok1[row]])

#с помощью for запишем наши данные
# - развернем в строку)
a = max_min_data('max_row')
for i in range (len(Spisok3)):
    _= active_sheet.cell(column= i + 1, row=a + 1, value=Spisok3[i])


#запишем данные в столбец после max_column - данных
b = max_min_data('max_column')
for i in range (len(Spisok2)):
    _= active_sheet.cell(column= b + 1, row=i + 1, value=Spisok2[i])

#запишем 2 списка сразу в 2 колонки

active_sheet2 = active_excel.create_sheet(title='MyGameDesign_2') #создали 2 лист в документе

Spisok4 = [
    1,
    2,
    3,
    4,
    5
]
Spisok5 = [
    6,
    7,
    8,
    9,
    10
]
Spisok6 = [
    11,
    12,
    13,
    14,
    15
]
Spisok7 = [
    16,
    17,
    18,
    19,
    20
]
Spisok8 = [
    21,
    22,
    23,
    24,
    25
]
S_45 = [Spisok4, Spisok5, Spisok6, Spisok7, Spisok8]

for i in range (len(Spisok4)):
    for j in range (len(S_45)):
        _=active_sheet2.cell(column=j+2, row=i+2, value=S_45[j][i])

#применение форматирование
style_1 = Font(name='Calibri', color=colors.DARKBLUE,
               bold=True, size=12)#underline='double'
style_2 = Font(name='Calibri', color=colors.YELLOW,
               bold=True, size=10, underline='single')
style_3 = Font(name='Calibri', color=colors.BLUE,
               bold=True, size=12, underline='double')
style_4 = Font(name='Calibri', color=colors.BLACK,
               bold=True, size=10)

#отформатируем шапку
a1 = active_sheet2['A1']
b1 = active_sheet2['B1']
a1.font = style_1
b1.font = style_1

#далее отформатируем остальную часть таблицы
for i in range(2,7):
    a = active_sheet2['B' + str(i)]
    b = active_sheet2['C' + str(i)]
    c = active_sheet2['D' + str(i)]
    d = active_sheet2['E' + str(i)]
    e = active_sheet2['F' + str(i)]

    a.font = style_4
    b.font = style_4
    c.font = style_4
    d.font = style_4
    e.font = style_4

#применем формулу СУММА
active_sheet2["A7"] = 'СУММА:'
#active_sheet2["B7"] = '=SUM(B1:B6)'
#active_sheet2["C7"] = '=SUM(C1:C6)'

a6 = active_sheet2['A7']
a6.font = style_3

#применим формулу макс - мин
active_sheet2["A8"] = 'МАКС:'
# active_sheet2["B8"] = '=max(B1:B6)'
# active_sheet2["C8"] = '=max(C1:C6)'

a7 = active_sheet2['A8']
a7.font = style_3

#применим формулу среднее значение
active_sheet2["A9"] = 'СР.ЗНАЧ:'
# active_sheet2["B9"] = '=average(B1:B6)'
# active_sheet2["C9"] = '=average(C1:C6)'

a8 = active_sheet2['A9']
a8.font = style_3

for i in range(2, 7):
    if i == 2:
        su = '=SUM(B1:B6)'
        ave = '=average(B1:B6)'
        mm = '=max(B1:B6)'
    elif i == 3:
        su = '=SUM(C1:C6)'
        ave = '=average(C1:C6)'
        mm = '=max(C1:C6)'
    elif i == 4:
        su = '=SUM(D1:D6)'
        ave = '=average(D1:D6)'
        mm = '=max(D1:D6)'
    elif i == 5:
        su = '=SUM(E1:E6)'
        ave = '=average(E1:E6)'
        mm = '=max(E1:E6)'
    elif i == 6:
        su = '=SUM(F1:F6)'
        ave = '=average(F1:F6)'
        mm = '=max(F1:F6)'

    _=active_sheet2.cell(column=i, row=(len(Spisok4)+2), value=su)
    _=active_sheet2.cell(column=i, row=(len(Spisok4)+4), value=ave)
    _.number_format='#,#0.0'
    _=active_sheet2.cell(column=i, row=(len(Spisok4)+3), value=mm)


for i in range(7,10):
    a = active_sheet2['B' + str(i)]
    b = active_sheet2['C' + str(i)]
    c = active_sheet2['D' + str(i)]
    d = active_sheet2['E' + str(i)]
    e = active_sheet2['F' + str(i)]

    a.font = style_1
    b.font = style_1
    c.font = style_1
    d.font = style_1
    e.font = style_1

active_excel.save('МойТест_3.xlsx') #сохраняем все изменения
print('ИЗМЕНЕНИЯ СОХРАНЕНЫ')
